import os
import openpyxl

def process_excel_files():
    """
    1. Удаляет столбцы по списку масок во всех файлах.
    2. Специально для листа 'Inputs':
       - Переименовывает столбец 'BlkLPT' в 'BlkLPT1'
       - Добавляет столбец 'BlkLPT2' (значение 0) справа от 'BlkLPT1'
    """
    # Список столбцов для удаления (по частичному совпадению)
    cols_to_delete = []
        #"T4_cbcswi1_swctrl", "T3_cbcswi1_swctrl", 
        #"SGF6_xcbr1_tsd", "T1_cbcswi1_hvbctrl", "distanz"
    #]
    
    # Настройки для специального листа Inputs
    target_sheet_name = "Inputs"
    
    # Настройки для переименования BlkLPT и добавления BlkLPT2
    old_col_name = "BlkLPT"
    new_col_name1 = "BlkLPT1"
    new_col_name2 = "BlkLPT2"
    new_col_value2 = 0

    current_dir = os.getcwd()
    xlsx_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    print(f"Найдено {len(xlsx_files)} xlsx файлов")
    print("-" * 50)
    
    for file_name in xlsx_files:
        file_path = os.path.join(current_dir, file_name)
        print(f"\nОбработка: {file_name}")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            changes_made = False
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = []
                
                # Сбор заголовков
                for col_idx in range(1, sheet.max_column + 1):
                    header = sheet.cell(row=1, column=col_idx).value
                    headers.append((col_idx, str(header) if header else ""))
                
                # --- СПЕЦИАЛЬНАЯ ЛОГИКА ДЛЯ ЛИСТА Inputs ---
                if sheet_name == target_sheet_name:
                    idx_blklpt = None
                    
                    # Поиск индекса столбца BlkLPT
                    for col_idx, header in headers:
                        if header == old_col_name:
                            idx_blklpt = col_idx
                            break

                    # 1. Переименование столбца 'BlkLPT' в 'BlkLPT1'
                    if idx_blklpt:
                        sheet.cell(row=1, column=idx_blklpt, value=new_col_name1)
                        print(f"  Лист '{sheet_name}': столбец '{old_col_name}' переименован в '{new_col_name1}'")
                        changes_made = True
                        
                        # 2. Добавление столбца 'BlkLPT2' справа от 'BlkLPT1'
                        insert_index = idx_blklpt + 1
                        sheet.insert_cols(insert_index)
                        cell = sheet.cell(row=1, column=insert_index, value=new_col_name2)
                        cell.font = openpyxl.styles.Font(bold=True)
                        
                        # Заполнение значением 0 (начиная со 2-й строки)
                        for row in range(2, sheet.max_row + 1):
                            sheet.cell(row=row, column=insert_index, value=new_col_value2)
                            
                        print(f"  Лист '{sheet_name}': добавлен столбец '{new_col_name2}' справа от '{new_col_name1}'")
                        changes_made = True
                    else:
                        print(f"  Лист '{sheet_name}': столбец '{old_col_name}' не найден")

                # --- ОБЩАЯ ЛОГИКА УДАЛЕНИЯ ПО МАСКАМ ---
                # Собираем заново заголовки, если лист был изменен (для Inputs)
                if sheet_name == target_sheet_name and changes_made:
                    # Обновляем список заголовков после манипуляций с Inputs
                    headers = []
                    for col_idx in range(1, sheet.max_column + 1):
                        header = sheet.cell(row=1, column=col_idx).value
                        headers.append((col_idx, str(header) if header else ""))

                cols_to_remove = []
                for col_idx, header in reversed(headers):
                    if header:
                        for pattern in cols_to_delete:
                            if pattern in header:
                                # Проверка: не удаляем новые столбцы BlkLPT1 и BlkLPT2
                                if sheet_name == target_sheet_name and (header == new_col_name1 or header == new_col_name2):
                                    continue
                                cols_to_remove.append(col_idx)
                                print(f"  Лист '{sheet_name}': удален столбец {col_idx} ('{header}') по маске '{pattern}'")
                                break
                
                # Удаление столбцов по маске
                if cols_to_remove:
                    for col_idx in sorted(cols_to_remove, reverse=True):
                        sheet.delete_cols(col_idx)
                    changes_made = True
            
            # Сохранение
            if changes_made:
                wb.save(file_path)
                print(f"  ✅ Изменения сохранены")
            else:
                print(f"  ℹ️  Изменений не требуется")
            
            wb.close()
            
        except Exception as e:
            print(f"  ❌ Ошибка: {e}")
    
    print("\n" + "-" * 50)
    print("Готово!")
    input("Нажмите Enter, чтобы выйти...")

if __name__ == "__main__":
    process_excel_files()